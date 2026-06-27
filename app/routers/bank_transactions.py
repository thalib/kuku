import json
import os
from calendar import month_abbr
from datetime import date

from fastapi import APIRouter, Request, Form, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import ValidationError

from app.config import APP_NAME, NAV_GROUPS
from app.database import get_db
from app.services import bank_accounts as bank_svc
from app.services import transactions as tx_svc
from app.models.transactions import TransactionUpdate
from app.utils.nav import mark_active_nav

router = APIRouter()

_template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates")
templates = Jinja2Templates(directory=_template_dir)

MONTHS = [
    (1, "JAN"), (2, "FEB"), (3, "MAR"), (4, "APR"),
    (5, "MAY"), (6, "JUN"), (7, "JUL"), (8, "AUG"),
    (9, "SEP"), (10, "OCT"), (11, "NOV"), (12, "DEC"),
]


def _base_ctx(request: Request, page_url: str, page_name: str):
    return {
        "app_name": APP_NAME,
        "page_title": f"{page_name} - Kuku",
        "nav_groups": mark_active_nav(NAV_GROUPS, page_url),
    }


@router.get("/transactions", response_class=HTMLResponse)
async def transactions_page(request: Request):
    db = await get_db()
    accounts = await bank_svc.list_accounts(db)
    active_accounts = [a for a in accounts if a["is_active"]]
    ctx = _base_ctx(request, "/banks/transactions", "Transactions")
    ctx["accounts"] = active_accounts
    return templates.TemplateResponse(request, "pages/banks_transactions.html", ctx)


@router.get("/transactions/filters", response_class=HTMLResponse)
async def transactions_filters(request: Request, account_id: int, selected_year: int | None = None):
    db = await get_db()
    years = await tx_svc.get_available_years(db, account_id)
    if not years:
        ctx = {"no_data": True, "account_id": account_id}
        return templates.TemplateResponse(request, "partials/tx_filters.html", ctx)
    yr = selected_year if selected_year and selected_year in years else max(years)
    avail = await tx_svc.get_available_months(db, account_id, yr)
    month_map = dict(MONTHS)
    months = [(m, month_map[m]) for m in avail]
    ctx = {"no_data": False, "account_id": account_id, "years": years,
           "months": months, "selected_year": yr, "selected_month": min(avail)}
    return templates.TemplateResponse(request, "partials/tx_filters.html", ctx)


@router.get("/transactions/table", response_class=HTMLResponse)
async def transactions_table(request: Request, account_id: int, year: int, month: int):
    db = await get_db()
    txns = await tx_svc.list_transactions(db, account_id, year, month)
    summary = await tx_svc.get_summary(db, account_id, year, month)
    account = await bank_svc.get_account(db, account_id)
    month_label = f"{month:02d} - {month_abbr[month]}"
    ctx = {
        "transactions": txns,
        "summary": summary,
        "account": account,
        "account_id": account_id,
        "year": year,
        "month": month,
        "month_label": month_label,
    }
    return templates.TemplateResponse(request, "partials/tx_table.html", ctx)


@router.get("/transactions/import/form", response_class=HTMLResponse)
async def import_form(request: Request, account_id: int):
    return templates.TemplateResponse(
        request, "partials/tx_import_form.html", {"account_id": account_id}
    )


@router.post("/transactions/import/preview", response_class=HTMLResponse)
async def import_preview(request: Request, account_id: int = Form(...), file: UploadFile = File(...)):
    filename = file.filename or "unknown"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    try:
        if ext == "csv":
            content = (await file.read()).decode("utf-8-sig")
            txns = tx_svc.parse_csv_rows(content)
        elif ext in ("xls", "xlsx"):
            data = await file.read()
            txns = tx_svc.parse_excel_bytes(data, filename)
        else:
            content = (await file.read()).decode("utf-8-sig")
            txns = tx_svc.parse_csv_rows(content)
    except Exception as e:
        return templates.TemplateResponse(
            request, "partials/tx_import_preview.html",
            {"error": f"Failed to parse file: {e}", "transactions": [], "account_id": account_id, "count": 0},
        )

    if not txns:
        return templates.TemplateResponse(
            request, "partials/tx_import_preview.html",
            {"error": "No valid transactions found in file.", "transactions": [], "account_id": account_id, "count": 0},
        )

    return templates.TemplateResponse(
        request, "partials/tx_import_preview.html",
        {"transactions": txns, "account_id": account_id, "error": None, "count": len(txns)},
    )


@router.post("/transactions/import/confirm", response_class=HTMLResponse)
async def import_confirm(request: Request, account_id: int = Form(...), data: str = Form(...)):
    db = await get_db()
    try:
        txns = json.loads(data)
    except json.JSONDecodeError:
        raise HTTPException(400, "Invalid transaction data")
    count = await tx_svc.bulk_create_transactions(db, account_id, txns)

    dates = [t["txn_date"] for t in txns if t.get("txn_date")]
    if dates:
        latest_date = max(dates)
        d = date.fromisoformat(latest_date)
        latest_year, latest_month = d.year, d.month
    else:
        latest_year, latest_month = date.today().year, date.today().month

    txns_list = await tx_svc.list_transactions(db, account_id, latest_year, latest_month)
    summary = await tx_svc.get_summary(db, account_id, latest_year, latest_month)
    account = await bank_svc.get_account(db, account_id)
    month_label = f"{latest_month:02d} - {month_abbr[latest_month]}"
    ctx = {
        "transactions": txns_list,
        "summary": summary,
        "account": account,
        "account_id": account_id,
        "year": latest_year,
        "month": latest_month,
        "month_label": month_label,
        "import_count": count,
    }
    resp = templates.TemplateResponse(request, "partials/tx_table.html", ctx)
    resp.headers["HX-Trigger"] = "txImported"
    return resp


@router.get("/transactions/{txn_id}/edit", response_class=HTMLResponse)
async def transaction_edit_form(request: Request, txn_id: int):
    db = await get_db()
    txn = await tx_svc.get_transaction(db, txn_id)
    if not txn:
        raise HTTPException(404)
    return templates.TemplateResponse(
        request, "partials/tx_edit_form.html", {"txn": txn}
    )


@router.get("/transactions/{txn_id}/cancel", response_class=HTMLResponse)
async def transaction_cancel_edit(request: Request, txn_id: int):
    db = await get_db()
    txn = await tx_svc.get_transaction(db, txn_id)
    if not txn:
        raise HTTPException(404)
    return templates.TemplateResponse(
        request, "partials/tx_row.html", {"txn": txn}
    )


@router.post("/transactions/{txn_id}/update", response_class=HTMLResponse)
async def transaction_update(
    request: Request,
    txn_id: int,
    txn_date: str = Form(...),
    value_date: str = Form(...),
    narration: str = Form(""),
    reference: str = Form(""),
    debit: str = Form("0"),
    credit: str = Form("0"),
    balance: str = Form("0"),
):
    db = await get_db()
    try:
        update_data = TransactionUpdate(
            txn_date=date.fromisoformat(txn_date),
            value_date=date.fromisoformat(value_date),
            narration=narration or None,
            reference=reference or None,
            debit=float(debit) if debit else 0,
            credit=float(credit) if credit else 0,
            balance=float(balance) if balance else 0,
        )
    except (ValueError, ValidationError):
        raise HTTPException(400, "Invalid data")
    updated = await tx_svc.update_transaction(db, txn_id, update_data)
    if not updated:
        raise HTTPException(404)
    txn = await tx_svc.get_transaction(db, txn_id)
    return templates.TemplateResponse(
        request, "partials/tx_row.html", {"txn": txn}
    )


@router.delete("/transactions/{txn_id}", response_class=HTMLResponse)
async def transaction_delete(request: Request, txn_id: int):
    db = await get_db()
    deleted = await tx_svc.delete_transaction(db, txn_id)
    if not deleted:
        raise HTTPException(404)
    return HTMLResponse("", status_code=200, headers={"HX-Trigger": "txDeleted"})


@router.get("/transactions/export/csv")
async def export_csv(account_id: int, year: int, month: int):
    import io, csv as csv_mod
    db = await get_db()
    txns = await tx_svc.list_transactions(db, account_id, year, month)
    summary = await tx_svc.get_summary(db, account_id, year, month)

    output = io.StringIO()
    writer = csv_mod.writer(output)
    writer.writerow(["Date", "Value Date", "Narration", "Reference", "Debit", "Credit", "Balance"])
    for t in txns:
        writer.writerow([t["txn_date"], t["value_date"], t["narration"], t["reference"],
                         t["debit"], t["credit"], t["balance"]])
    writer.writerow([])
    writer.writerow(["", "", "Totals", "", summary["total_debit"], summary["total_credit"], ""])

    output.seek(0)
    content = output.getvalue().encode("utf-8-sig")
    fname = f"transactions_{account_id}_{year}_{month:02d}.csv"
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/transactions/export/xlsx")
async def export_xlsx(account_id: int, year: int, month: int):
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment

    db = await get_db()
    txns = await tx_svc.list_transactions(db, account_id, year, month)
    summary = await tx_svc.get_summary(db, account_id, year, month)
    account = await bank_svc.get_account(db, account_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws.append(["Transactions"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    acct_label = f"{account['bank_name']} - ****{account['account_number'][-4:]}" if account else f"Account #{account_id}"
    ws.append([acct_label, f"{year}-{month:02d}"])
    ws.cell(row=2, column=1).font = Font(italic=True, size=11)
    ws.append([])

    ws.append(["Date", "Value Date", "Narration", "Reference", "Debit", "Credit", "Balance"])
    for c in ws[4]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for t in txns:
        ws.append([t["txn_date"], t["value_date"], t["narration"], t["reference"],
                   t["debit"], t["credit"], t["balance"]])

    last_data_row = 4 + len(txns)
    ws.append([])
    ws.append(["", "", "Total Debit", summary["total_debit"], "Total Credit", summary["total_credit"], ""])

    for col_idx in (5, 6, 7):
        for row_idx in range(5, last_data_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"
    for col_idx in (4, 6):
        ws.cell(row=last_data_row + 2, column=col_idx).number_format = "#,##0.00"
        ws.cell(row=last_data_row + 2, column=col_idx).font = Font(bold=True)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"transactions_{account_id}_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/transactions/export/pdf")
async def export_pdf(account_id: int, year: int, month: int):
    import io
    from fpdf import FPDF

    db = await get_db()
    txns = await tx_svc.list_transactions(db, account_id, year, month)
    summary = await tx_svc.get_summary(db, account_id, year, month)
    account = await bank_svc.get_account(db, account_id)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Bank Transactions", new_x="LMARGIN", new_y="NEXT", align="C")

    pdf.set_font("Helvetica", "", 11)
    if account:
        acct_label = f"{account['bank_name']} - ****{account['account_number'][-4:]}"
    else:
        acct_label = f"Account #{account_id}"
    month_label = f"{month:02d} - {month_abbr[month]}"
    pdf.cell(0, 7, f"{acct_label}  |  {year}-{month_label}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 9)
    col_w = [22, 75, 25, 22, 22, 22]
    hdrs = ["Date", "Narration", "Ref", "Debit", "Credit", "Balance"]
    for i, h in enumerate(hdrs):
        pdf.cell(col_w[i], 7, h, border=1, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    for t in txns:
        narr = t["narration"] or ""
        if len(narr) > 50:
            narr = narr[:47] + "..."
        rh = 6
        x0 = pdf.get_x()
        y0 = pdf.get_y()
        pdf.rect(x0, y0, col_w[0], rh)
        pdf.cell(col_w[0], rh, str(t["txn_date"]), align="C")
        pdf.set_xy(x0 + col_w[0], y0)
        pdf.cell(col_w[1], rh, narr)
        pdf.set_xy(x0 + sum(col_w[:2]), y0)
        pdf.cell(col_w[2], rh, str(t["reference"] or ""))
        pdf.set_xy(x0 + sum(col_w[:3]), y0)
        pdf.cell(col_w[3], rh, f"{t['debit']:,.2f}" if t["debit"] else "", align="R")
        pdf.set_xy(x0 + sum(col_w[:4]), y0)
        pdf.cell(col_w[4], rh, f"{t['credit']:,.2f}" if t["credit"] else "", align="R")
        pdf.set_xy(x0 + sum(col_w[:5]), y0)
        pdf.cell(col_w[5], rh, f"{t['balance']:,.2f}", align="R")
        pdf.ln()

    pdf.ln(3)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, f"Total Debit: {summary['total_debit']:,.2f}   |   Total Credit: {summary['total_credit']:,.2f}   |   Transactions: {summary['txn_count']}", new_x="LMARGIN", new_y="NEXT", align="C")

    buf = io.BytesIO()
    buf.write(pdf.output())
    buf.seek(0)
    fname = f"transactions_{account_id}_{year}_{month:02d}.pdf"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
