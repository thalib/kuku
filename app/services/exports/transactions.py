import csv
import io

import openpyxl
from openpyxl.styles import Alignment, Font
from xhtml2pdf import pisa

from app.services.exports.base import BaseExporter, build_month_label, build_category_display, get_template


class TransactionExporter(BaseExporter):
    domain = "transactions"

    def __init__(self, account: dict | None = None, account_id: int = 0,
                 fy: int = 0, calendar_year: int = 0, month: int = 0):
        super().__init__(account, account_id)
        self.fy = fy
        self.calendar_year = calendar_year
        self.month = month
        self.month_label = build_month_label(calendar_year, month)

    def render_csv(self, transactions: list[dict], summary: dict) -> bytes:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([self.company_name])
        writer.writerow([self.account_label])
        writer.writerow([self.month_label])
        writer.writerow([])
        writer.writerow(["Date", "Narration", "Category", "Debit", "Credit", "Balance"])
        for t in transactions:
            narration = t["narration"] or ""
            if t.get("reference"):
                narration += f"\n({t['reference']})"
            category = build_category_display(t)
            writer.writerow([
                t["txn_date"], narration, category,
                t["debit"], t["credit"], t["balance"],
            ])
        writer.writerow([])
        writer.writerow(["", "Totals", "", summary["total_debit"], summary["total_credit"], ""])
        output.seek(0)
        return output.getvalue().encode("utf-8-sig")

    def render_xlsx(self, transactions: list[dict], summary: dict) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        ws.append([self.company_name])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([self.account_label])
        ws.cell(row=2, column=1).font = Font(italic=True, size=11)
        ws.append([self.month_label])
        ws.append([])

        headers = ["Date", "Narration", "Category", "Debit", "Credit", "Balance"]
        ws.append(headers)
        header_row = ws[5]
        for c in header_row:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        for t in transactions:
            narration = t["narration"] or ""
            if t.get("reference"):
                narration += f"\n({t['reference']})"
            category = build_category_display(t)
            row = [t["txn_date"], narration, category, t["debit"], t["credit"], t["balance"]]
            ws.append(row)

        data_start_row = 6
        for row_idx in range(data_start_row, data_start_row + len(transactions)):
            for col_idx in (4, 5, 6):
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"
            if transactions[row_idx - data_start_row].get("debit"):
                ws.cell(row=row_idx, column=4).font = Font(color="cc0000")
            if transactions[row_idx - data_start_row].get("credit"):
                ws.cell(row=row_idx, column=5).font = Font(color="007700")
            narration_cell = ws.cell(row=row_idx, column=2)
            narration_cell.alignment = Alignment(wrap_text=True)

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12

        summary_row = data_start_row + len(transactions)
        ws.append([])
        ws.append(["", "Totals", "", summary["total_debit"], summary["total_credit"], ""])
        total_row = summary_row + 2
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        ws.cell(row=total_row, column=4).number_format = "#,##0.00"
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        ws.cell(row=total_row, column=5).number_format = "#,##0.00"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def render_pdf(self, transactions: list[dict], summary: dict) -> bytes:
        template = get_template(self.domain, "pdf.html")
        html_content = template.render(
            **self.base_context,
            bank_name=self.account_label,
            month_label=self.month_label,
            transactions=transactions,
            summary=summary,
            format_category=build_category_display,
        )

        pdf_bytes = io.BytesIO()
        pisa_status = pisa.CreatePDF(io.StringIO(html_content), dest=pdf_bytes)
        if pisa_status.err:
            raise RuntimeError(f"PDF generation failed with {pisa_status.err} errors")
        pdf_bytes.seek(0)
        return pdf_bytes.getvalue()
